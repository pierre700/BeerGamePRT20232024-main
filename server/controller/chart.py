import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def afficher_graphiques_excel(nom_fichier_excel, feuille):
    # Lecture du fichier Excel
    try:
        if feuille:
            df = pd.read_excel(nom_fichier_excel, sheet_name=feuille)
        else:
            df = pd.read_excel(nom_fichier_excel)
    except FileNotFoundError:
        print("Fichier introuvable.")
        return
    except Exception as e:
        print("Erreur lors de la lecture du fichier Excel:", e)
        return
    
    # Vérifier si le DataFrame est vide
    if df.empty:
        print("Le DataFrame est vide. Aucune donnée à afficher.")
        return
    
    # Sélection de la première colonne comme axe x
    colonne_x = df.columns[0]
    
    # Sélection des autres colonnes comme axes y
    colonnes_y = df.columns[1:4]
    
    # Affichage des graphiques
    for colonne_y in colonnes_y:
        plt.figure(figsize=(10, 6))
        plt.plot(df[colonne_x], df[colonne_y], marker='o', linestyle='-')
        plt.title(f"{colonne_y} en fonction de {colonne_x} pour le {feuille}")
        plt.xlabel(colonne_x)
        plt.ylabel(colonne_y)
        plt.grid(True)
        plt.show()


def afficher_graphiques_excel_toutes_feuilles():
    # Chemin d'accès au fichier Excel (chemin relatif)
    chemin_fichier_excel = "export.xlsx"
    
    # Lecture du fichier Excel et récupération des noms de toutes les feuilles
    try:
        workbook = load_workbook(chemin_fichier_excel)
        feuilles = workbook.sheetnames
    except FileNotFoundError:
        print("Fichier introuvable.")
        return
    except Exception as e:
        print("Erreur lors de la lecture du fichier Excel:", e)
        return
    
    # Affichage des graphiques pour chaque feuille
    for feuille in feuilles:
        afficher_graphiques_excel(chemin_fichier_excel, feuille)


# Exemple d'utilisation
afficher_graphiques_excel_toutes_feuilles()